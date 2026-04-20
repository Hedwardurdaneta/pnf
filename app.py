import flet as ft
import gspread
import openpyxl
import os
import random
import asyncio
import json
from google.oauth2.service_account import Credentials

# --- [ CONFIGURACION ] ---
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
CREDS_JSON = os.path.join(BASE_DIR, "credentials.json")
EXCEL_LOCAL = os.path.join(BASE_DIR, "Programacion.xlsx")
COLOR_FONDO = "#0c6980"
COLOR_BOTON = "#f0f4fa"
COLOR_TEXTO_BOTON = "#1976d2"

# --- [ BANCO DE DATOS ] ---
CONTENIDO = {
    "UNIDAD I": {
        "Algoritmo": "Secuencia de pasos logicos para resolver un problema.",
        "IDE": "Entorno de Desarrollo Integrado para escribir codigo.",
        "Depuracion": "Proceso de identificar y corregir errores en el codigo.",
        "Compilacion": "Traduccion de codigo de alto nivel a lenguaje maquina.",
        "Sintaxis": "Reglas que definen como escribir instrucciones.",
        "Variable": "Espacio en memoria para almacenar un dato.",
        "Codigo Fuente": "Instrucciones escritas por el programador.",
        "Comentario": "Lineas ignoradas por el compilador para documentar.",
        "Hardware": "Componentes fisicos del sistema informatico.",
        "Software": "Programas y reglas logicas del sistema."
    },
    "UNIDAD II": {
        "int": "Tipo de dato para numeros enteros.",
        "float": "Tipo de dato para numeros decimales.",
        "str": "Cadenas de texto o caracteres.",
        "bool": "Tipo logico: True (Verdadero) o False (Falso).",
        "Lista": "Coleccion organizada de multiples valores.",
        "Operador": "Simbolos para realizar operaciones (+, -, *, /).",
        "Asignacion": "Guardar un valor en una variable usando '='.",
        "if": "Condicional que ejecuta codigo si se cumple algo.",
        "while": "Bucle que repite codigo mientras se cumpla una condicion.",
        "for": "Bucle para repetir codigo un numero fijo de veces."
    },
    "UNIDAD III": {
        "Flet": "Framework para crear interfaces con Python.",
        "Widget": "Componente visual basico (boton, imagen, etc.).",
        "Label": "Control para mostrar texto estatico.",
        "Entry": "Campo de texto para entrada del usuario.",
        "Button": "Componente interactivo para ejecutar acciones.",
        "Container": "Agrupador de elementos con estilo.",
        "Evento": "Accion detectada como un clic o tecla pulsada.",
        "Layout": "Organizacion visual de los elementos.",
        "Mainloop": "Bucle que mantiene la app abierta e interactiva.",
        "Color": "Atributo para personalizar fondos y textos."
    }
}

PREGUNTAS = {
    "UNIDAD I": [
        ("Que es un algoritmo?", ["Pasos logicos", "Un virus", "Hardware"], "Pasos logicos"),
        ("Que significa IDE?", ["Entorno de Desarrollo", "Internet", "Disco"], "Entorno de Desarrollo"),
        ("Que es la depuracion?", ["Corregir errores", "Borrar archivos", "Instalar"], "Corregir errores"),
        ("Que hace la compilacion?", ["Traducir codigo", "Apagar PC", "Imprimir"], "Traducir codigo"),
        ("Que es la sintaxis?", ["Reglas de escritura", "Teclado", "Monitor"], "Reglas de escritura"),
        ("Donde se guarda una variable?", ["Memoria", "Caja", "Papel"], "Memoria"),
        ("Que es el codigo fuente?", ["Instrucciones", "Electricidad", "Agua"], "Instrucciones"),
        ("El compilador lee comentarios?", ["No", "Si", "A veces"], "No"),
        ("Que es el hardware?", ["Parte fisica", "Programas", "Internet"], "Parte fisica"),
        ("Que es el software?", ["Parte logica", "Monitor", "Cables"], "Parte logica")
    ],
    "UNIDAD II": [
        ("Que guarda un 'int'?", ["Enteros", "Letras", "Imagenes"], "Enteros"),
        ("Que guarda un 'float'?", ["Decimales", "Cadenas", "Enteros"], "Decimales"),
        ("Que es un 'str'?", ["Texto", "Numeros", "Bucle"], "Texto"),
        ("Valores del 'bool'?", ["True/False", "A/B", "1/100"], "True/False"),
        ("Que es una lista?", ["Coleccion", "Variable unica", "Error"], "Coleccion"),
        ("Que es '+'?", ["Operador", "Variable", "Widget"], "Operador"),
        ("Simbolo de asignacion?", ["=", "==", "+"], "="),
        ("Que es 'if'?", ["Condicional", "Bucle", "Variable"], "Condicional"),
        ("Que es 'while'?", ["Bucle", "Salida", "Suma"], "Bucle"),
        ("Que es 'for'?", ["Bucle repetitivo", "Suma", "Texto"], "Bucle repetitivo")
    ],
    "UNIDAD III": [
        ("Para que sirve Flet?", ["Interfaces", "Hardware", "Cafe"], "Interfaces"),
        ("Que es un Widget?", ["Componente visual", "Cable", "Virus"], "Componente visual"),
        ("Que muestra un Label?", ["Texto", "Video", "Musica"], "Texto"),
        ("Que es un Entry?", ["Entrada de texto", "Salida", "Imagen"], "Entrada de texto"),
        ("Que hace un Button?", ["Ejecuta accion", "Nada", "Cierra"], "Ejecuta accion"),
        ("Que es un Container?", ["Agrupador", "Variable", "Lista"], "Agrupador"),
        ("Que es un clic?", ["Evento", "Error", "Hardware"], "Evento"),
        ("Que es el Layout?", ["Organizacion", "Color", "Nombre"], "Organizacion"),
        ("Que es el Mainloop?", ["Bucle de la app", "Cable", "Boton"], "Bucle de la app"),
        ("El color es un atributo?", ["Si", "No", "Solo web"], "Si")
    ]
}

# --- [ MOTOR DE NUBE ] ---
class CloudService:
    def __init__(self):
        self.sheet = None
        self.last_error = ""
        self.creds_source = "Ninguna"
        self._connect()

    def _connect(self):
        try:
            scopes = [
                "https://www.googleapis.com/auth/spreadsheets",
                "https://www.googleapis.com/auth/drive"
            ]
            
            creds = None
            
            google_creds_env = os.getenv("GOOGLE_CREDENTIALS")
            if google_creds_env:
                try:
                    creds_info = json.loads(google_creds_env)
                    creds = Credentials.from_service_account_info(creds_info, scopes=scopes)
                    self.creds_source = "Variable de entorno"
                except Exception as e:
                    self.last_error = f"Error con variable de entorno: {e}"
                    return
            
            elif os.path.exists(CREDS_JSON):
                try:
                    creds = Credentials.from_service_account_file(CREDS_JSON, scopes=scopes)
                    self.creds_source = "Archivo local"
                except Exception as e:
                    self.last_error = f"Error leyendo archivo: {e}"
                    return
            
            else:
                self.last_error = "No se encontraron credenciales"
                return
            
            if not creds:
                return
            
            client = gspread.authorize(creds)
            workbook = client.open("Ingenieria de software II")
            self.sheet = workbook.worksheet("Notas_PNF_UNERMB")
            
        except Exception as e:
            self.last_error = f"Error general: {e}"

    def update_nota(self, cedula, unidad, nota):
        if not self.sheet:
            return False
        
        try:
            ceds = self.sheet.col_values(2)
            cedula_str = str(cedula).strip()
            
            if cedula_str not in ceds:
                return False
            
            row = ceds.index(cedula_str) + 1
            col = {"UNIDAD I": 4, "UNIDAD II": 5, "UNIDAD III": 6}.get(unidad, 4)
            
            self.sheet.update_cell(row, col, nota)
            return True
            
        except Exception as e:
            self.last_error = f"Error al actualizar: {e}"
            return False

    def verificar_nota_existente(self, cedula, unidad):
        """Verifica si el alumno ya tiene una nota registrada"""
        if not self.sheet:
            return False  # Si no hay conexion, permitir intento (fallback)
        
        try:
            ceds = self.sheet.col_values(2)
            cedula_str = str(cedula).strip()
            
            if cedula_str not in ceds:
                return False  # No encontrado, no ha presentado
            
            row = ceds.index(cedula_str) + 1
            col = {"UNIDAD I": 4, "UNIDAD II": 5, "UNIDAD III": 6}.get(unidad, 4)
            
            # Obtener el valor actual de la celda
            nota_actual = self.sheet.cell(row, col).value
            
            # Si la celda tiene algun valor (no esta vacia), ya presento
            if nota_actual is not None and str(nota_actual).strip() != "":
                return True  # Ya tiene nota
            
            return False  # No tiene nota, puede presentar
            
        except Exception as e:
            self.last_error = f"Error al verificar: {e}"
            return False  # En caso de error, permitir el intento

# --- [ APLICACION FLET ] ---
async def main(page: ft.Page):
    page.title = "SISTEMA ACADEMICO UNERMB"
    page.bgcolor = COLOR_FONDO
    page.horizontal_alignment = ft.CrossAxisAlignment.CENTER
    
    cloud = CloudService()
    state = {"name": "", "id": "", "unit": "", "pts": 0, "idx": 0, "timer_active": False}

    async def navigate(func):
        page.clean()
        await func()
        page.update()

    async def view_login():
        students = {}
        if os.path.exists(EXCEL_LOCAL):
            wb = openpyxl.load_workbook(EXCEL_LOCAL, data_only=True)
            ws = wb.active
            for r in range(2, 100):
                c, n = ws.cell(r, 2).value, ws.cell(r, 3).value
                if n: students[str(n)] = str(c)

        dd = ft.Dropdown(label="Seleccione su Nombre", width=400, bgcolor="white",
                         options=[ft.dropdown.Option(s) for s in students.keys()])

        tf = ft.TextField(label="Cedula", password=True, width=400, bgcolor="white")

        async def do_login(e):
            if dd.value and students.get(dd.value) == tf.value:
                state["name"], state["id"] = dd.value, tf.value
                await navigate(view_menu)
            else:
                page.snack_bar = ft.SnackBar(ft.Text("Credenciales Incorrectas"), bgcolor="red")
                page.snack_bar.open = True
                page.update()

        page.add(
            ft.Text("INGENIERIA DE SOFTWARE II", size=30, weight="bold", color="white"),
            ft.Container(height=20),
            ft.Container(content=ft.Column([dd, tf, ft.ElevatedButton("ENTRAR", on_click=do_login, width=200)],
                                          horizontal_alignment="center"), 
                         padding=40, bgcolor="#33ffffff", border_radius=20)
        )

    async def view_menu():
        page.add(
            ft.Text(f"Bienvenido: {state['name']}", size=20, color="white"),
            ft.Divider(color="white"),
            *[ft.ElevatedButton(u, on_click=lambda e, x=u: asyncio.run(start_unit(x)), width=350, height=55) for u in CONTENIDO.keys()],
            ft.TextButton("Cerrar Sesion", on_click=lambda _: asyncio.run(navigate(view_login)), style=ft.ButtonStyle(color="white"))
        )

    async def start_unit(u):
        state["unit"] = u
        
        # VERIFICAR SI YA PRESENTO LA EVALUACION
        ya_presento = cloud.verificar_nota_existente(state["id"], state["unit"])
        
        if ya_presento:
            await navigate(lambda: view_ya_presento(u))
        else:
            await navigate(view_study)

    async def view_ya_presento(unidad):
        """Vista que muestra cuando el alumno ya presento la evaluacion"""
        page.add(
            ft.Icon(ft.icons.CHECK_CIRCLE, size=80, color="yellow"),
            ft.Container(height=20),
            ft.Text("EVALUACION YA REALIZADA", size=28, color="white", weight="bold"),
            ft.Container(height=10),
            ft.Text(f"Ya presentaste la evaluacion de:", size=18, color="white"),
            ft.Text(unidad, size=24, color="yellow", weight="bold"),
            ft.Container(height=10),
            ft.Text("Solo se permite una oportunidad por unidad.", size=14, color="white70"),
            ft.Container(height=30),
            ft.ElevatedButton(
                "REGRESAR AL MENU",
                on_click=lambda _: asyncio.run(navigate(view_menu)),
                width=300,
                height=50
            )
        )

    async def view_study():
        items = [ft.ListTile(title=ft.Text(k, weight="bold"), subtitle=ft.Text(v)) for k, v in CONTENIDO[state["unit"]].items()]
        page.add(
            ft.Text(f"MATERIAL: {state['unit']}", size=24, color="white", weight="bold"),
            ft.Container(content=ft.Column(items, scroll=ft.ScrollMode.ALWAYS, height=400), bgcolor="white", padding=15, border_radius=10, width=600),
            ft.Row([
                ft.ElevatedButton("EXAMEN", on_click=lambda _: asyncio.run(start_exam()), bgcolor="#0c6980", color="white"),
                ft.ElevatedButton("VOLVER", on_click=lambda _: asyncio.run(navigate(view_menu)))
            ], alignment="center")
        )

    async def start_exam():
        state["pts"], state["idx"] = 0, 0
        await navigate(view_exam)

    async def view_exam():
        bank = PREGUNTAS[state["unit"]]
        if state["idx"] < 10:
            q, opts, ans = bank[state["idx"]]
            opciones_mezcladas = list(opts)
            random.shuffle(opciones_mezcladas)

            lbl_timer = ft.Text("15", size=35, weight="bold", color="yellow")
            current_instance_idx = state["idx"]
            state["timer_active"] = True

            async def check(pick):
                if state["timer_active"]:
                    state["timer_active"] = False
                    if pick == ans: state["pts"] += 1
                    state["idx"] += 1
                    await navigate(view_exam)

            page.add(
                ft.Row([ft.Icon(ft.icons.TIMER, color="white"), lbl_timer], alignment="center"),
                ft.Text(f"Pregunta {state['idx']+1} de 10", color="white"),
                ft.Container(content=ft.Text(q, size=22, weight="bold", text_align="center"), padding=35, bgcolor="white", border_radius=15, width=650),
                *[ft.ElevatedButton(o, on_click=lambda e, x=o: asyncio.run(check(x)), width=450, height=50, 
                                   style=ft.ButtonStyle(bgcolor=COLOR_BOTON, color=COLOR_TEXTO_BOTON)) for o in opciones_mezcladas]
            )

            for i in range(15, -1, -1):
                if not state["timer_active"] or state["idx"] != current_instance_idx:
                    break
                lbl_timer.value = str(i)
                if i <= 5: lbl_timer.color = "red"
                page.update()
                await asyncio.sleep(1)
                
                if i == 0 and state["timer_active"] and state["idx"] == current_instance_idx:
                    state["timer_active"] = False
                    state["idx"] += 1
                    await navigate(view_exam)
        else:
            await navigate(view_result)

    async def view_result():
        page.add(ft.ProgressRing(), ft.Text("Guardando nota...", color="white"))
        page.update()
        
        success = cloud.update_nota(state["id"], state["unit"], state["pts"])
        
        page.clean()
        page.add(
            ft.Text("EVALUACION FINALIZADA", size=28, color="white", weight="bold"),
            ft.Text(f"{state['pts']}/10", size=110, color="yellow", weight="bold"),
            ft.Row([
                ft.Icon(
                    ft.icons.CLOUD_DONE if success else ft.icons.CLOUD_OFF, 
                    color="green" if success else "red"
                ),
                ft.Text(
                    "Nota guardada correctamente" if success else "Error al guardar la nota", 
                    color="white"
                )
            ], alignment="center"),
            ft.ElevatedButton(
                "REGRESAR AL MENU", 
                on_click=lambda _: asyncio.run(navigate(view_menu)), 
                width=300
            )
        )

    await view_login()

if __name__ == "__main__":
    ft.app(target=main, view=ft.AppView.WEB_BROWSER, port=int(os.getenv("PORT", 8080)), host="0.0.0.0")
